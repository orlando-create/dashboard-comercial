"""
ATUALIZAR DASHBOARD — Indicadores Time Comercial
=================================================
Execute este script sempre que atualizar a planilha Excel.
Ele lê a Base de Dados (incluindo a coluna ORIGEM) e regenera
os dados do One_Page_Comercial_2026.html automaticamente.

Uso: python atualizar_dashboard.py
"""

import openpyxl, json, re, unicodedata, os, sys
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, 'Dashboard_Comercial_2026.xlsx')
HTML_PATH  = os.path.join(BASE_DIR, 'One_Page_Comercial_2026.html')

# ── helpers ──────────────────────────────────────────────────────────────────

def parse_regiao(praca):
    if not praca: return 'OUTROS', 'OUTROS', 'OUTROS'
    p = str(praca).strip()
    mapping = {
        'São Paulo e Santos': ('SAO PAULO / SP', 'SP', 'SAO PAULO'),
        'Campinas':           ('CAMPINAS / SP',  'SP', 'CAMPINAS'),
        'Fortaleza':          ('FORTALEZA / CE', 'CE', 'FORTALEZA'),
        'Curitiba':           ('CURITIBA / PR',  'PR', 'CURITIBA'),
        'Goiânia':            ('GOIANIA / GO',   'GO', 'GOIANIA'),
        'Goiania':            ('GOIANIA / GO',   'GO', 'GOIANIA'),
        'Minas Gerais':       ('BELO HORIZONTE / MG', 'MG', 'BELO HORIZONTE'),
        'Rio de Janeiro':     ('RIO DE JANEIRO / RJ', 'RJ', 'RIO DE JANEIRO'),
        'Ribeirão Preto':     ('RIBEIRAO PRETO / SP', 'SP', 'RIBEIRAO PRETO'),
    }
    for k, v in mapping.items():
        if k.lower() in p.lower():
            return v
    uf = p.split('/')[-1].strip().upper() if '/' in p else 'OUTROS'
    return (p.upper(), uf, p.upper())

# ── leitura do Excel ──────────────────────────────────────────────────────────

print(f'[{datetime.now():%H:%M:%S}] Lendo {EXCEL_PATH}...')
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Base de Dados']

# Mapear cabeçalhos dinamicamente
headers = {ws.cell(1, c).value: c-1 for c in range(1, ws.max_column+1) if ws.cell(1, c).value}
print(f'  Colunas encontradas: {list(headers.keys())}')

def col(row, name, default=None):
    idx = headers.get(name)
    if idx is None: return default
    v = row[idx]
    return v if v is not None else default

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    hunter = col(row, 'HUNTER')
    if not hunter: continue
    regiao, uf, cidade = parse_regiao(col(row, 'PRAÇA'))
    parceiro = col(row, 'PARCEIRO')
    origem_raw = col(row, 'ORIGEM')
    origem = str(origem_raw).strip().upper() if origem_raw else None

    obj = {
        'hunter':      str(hunter).strip(),
        'empresa':     str(col(row,'EMPRESA','')).strip() or None,
        'regiao':      regiao,
        'uf':          uf,
        'cidade':      cidade,
        'produto':     str(col(row,'PRODUTO','')).strip() or None,
        'contrato':    str(col(row,'CONTRATO','NOVO')).strip(),
        'parceiro':    'SIM' if parceiro else 'NÃO',
        'honorarios':  round(float(col(row,'HONORÁRIOS',0)), 4),
        'faturamento': round(float(col(row,'FATURAMENTO',0)), 2),
        'credito':     round(float(col(row,'CRÉDITO',0)), 2),
        'status':      str(col(row,'STATUS','')).strip() or None,
        'mes':         str(col(row,'MÊS','')).strip() or None,
        'mes_assinado':str(col(row,'MÊS ASSINADO','')).strip() or None,
        'temperatura': str(col(row,'TEMPERATURA','')).strip() or None,
        'origem':      origem,
    }
    rows.append(obj)

total = len(rows)
apas_count = sum(1 for r in rows if r['origem'] and 'APAS' in r['origem'])
origens = sorted(set(r['origem'] for r in rows if r['origem']))
print(f'  Total registros: {total}')
print(f'  Registros APAS: {apas_count}')
print(f'  Origens únicas: {origens if origens else "(nenhuma preenchida ainda)"}')

# ── injetar no HTML ───────────────────────────────────────────────────────────

print(f'[{datetime.now():%H:%M:%S}] Atualizando {HTML_PATH}...')
with open(HTML_PATH, 'r', encoding='utf-8') as f:
    html = f.read()

new_raw = 'const RAW = ' + json.dumps(rows, ensure_ascii=False, separators=(',', ':')) + ';'
html_new = re.sub(r'const RAW = \[.*?\];', new_raw, html, flags=re.DOTALL)

if html_new == html:
    print('  AVISO: nada foi substituído — verifique se o HTML possui "const RAW = [...];"')
    sys.exit(1)

with open(HTML_PATH, 'w', encoding='utf-8') as f:
    f.write(html_new)

print(f'  ✓ Dashboard atualizado com {total} registros ({apas_count} da APAS).')
print(f'[{datetime.now():%H:%M:%S}] Concluído! Recarregue o One_Page_Comercial_2026.html no navegador.')
