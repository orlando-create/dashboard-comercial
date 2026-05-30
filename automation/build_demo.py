#!/usr/bin/env python3
"""
Gerador da DEMO multi-pagina do Dashboard Comercial 2026.
Le forescast atualizado 2026.xlsx, normaliza dados e produz:
  Indicadores Time Comercial/demo/One_Page_Comercial_2026_DEMO.html
"""
import os, sys, json, re, datetime, subprocess, shutil, tempfile
import openpyxl

def _detect_mnt():
    base = '/sessions'
    if os.path.isdir(base):
        for d in os.listdir(base):
            mnt = os.path.join(base, d, 'mnt')
            if os.path.isdir(mnt) and os.path.isdir(os.path.join(mnt, 'Indicadores Time Comercial')):
                return os.path.join(base, d, 'mnt')
    raise RuntimeError('mnt nao encontrado')

MNT = _detect_mnt()
SRC = os.path.join(MNT, 'Planejamento 2026', 'forescast atualizado 2026.xlsx')
OUT_DIR = os.path.join(MNT, 'Indicadores Time Comercial', 'demo')
OUT = os.path.join(OUT_DIR, 'One_Page_Comercial_2026_DEMO.html')
os.makedirs(OUT_DIR, exist_ok=True)

MONTHS = ['JAN','FEV','MAR','ABR','MAI','JUN','JUL','AGO','SET','OUT','NOV','DEZ']
MONTH_FIX = {'ABRI':'ABR','MAIO':'MAI','JANEIRO':'JAN','FEVEREIRO':'FEV','MARCO':'MAR','MARÇO':'MAR','JUNHO':'JUN','JULHO':'JUL','AGOSTO':'AGO','SETEMBRO':'SET','OUTUBRO':'OUT','NOVEMBRO':'NOV','DEZEMBRO':'DEZ','NAN':'','NONE':''}
HUNTER_FIX = {'NATALIA':'NATHALIA','THIAGO ':'THIAGO','KATERYNI':'KATARINY'}

def parse_brl(v):
    if v is None: return 0.0
    if isinstance(v,(int,float)): return float(v)
    s = str(v).strip().replace('R$','').replace(' ','').replace('.','').replace(',','.')
    try: return float(s)
    except: return 0.0
def get_uf(r):
    if not r: return 'N/D'
    s = str(r).strip().upper()
    m = re.search(r'-\s*([A-Z]{2})\s*$', s)
    if m: return m.group(1)
    if s == 'SAO PAULO': return 'SP'
    return 'N/D'
def clean_city(r):
    if not r: return 'N/D'
    s = str(r).strip().upper()
    s = re.sub(r'\s*-\s*[A-Z]{2}\s*$','', s).strip()
    s = re.sub(r'\s+',' ', s)
    return s if s else 'N/D'

def load_wb(path):
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception:
        tmp = tempfile.mkdtemp(prefix='xfix_')
        b = os.path.join(tmp,'b.xlsx'); r = os.path.join(tmp,'r.xlsx')
        shutil.copyfile(path,b)
        subprocess.run(['zip','-FF',b,'--out',r], input=b'y\n', capture_output=True, timeout=60)
        return openpyxl.load_workbook(r, data_only=True)

def main():
    wb = load_wb(SRC)
    sh = wb['FORECAST 2026']
    rows = list(sh.iter_rows(values_only=True))
    headers = list(rows[0])
    cols = {h:i for i,h in enumerate(headers) if h}
    records = []
    for row in rows[1:]:
        if not row or row[cols.get('HUNTER',0)] is None: continue
        hunter = str(row[cols['HUNTER']]).strip().upper()
        hunter = HUNTER_FIX.get(hunter, hunter)
        empresa = str(row[cols['EMPRESA']]).strip() if row[cols['EMPRESA']] else ''
        if not empresa: continue
        regiao = row[cols.get('REGIÃO',2)]
        produto = (row[cols.get('PRODUTO',3)] or 'OUTROS')
        contrato = (str(row[cols['CONTRATO']]).strip().upper() if row[cols.get('CONTRATO',4)] else 'NOVO')
        contrato = 'NOVO' if contrato in ('','NAN','NONE') else contrato
        parceiro = (str(row[cols['PARCEIRO']]).strip().upper() if row[cols.get('PARCEIRO',5)] else 'NÃO')
        honor = row[cols.get('HONORÁRIOS',7)] or 0
        fat = parse_brl(row[cols.get('FATURAMENTO',8)])
        cred = parse_brl(row[cols.get('CRÉDITO',6)])
        status = (str(row[cols['STATUS']]).strip().upper() if row[cols.get('STATUS',9)] else '')
        mes = (str(row[cols['MES']]).strip().upper() if row[cols.get('MES',10)] else '')
        mes = MONTH_FIX.get(mes, mes)
        mes_ass = (str(row[cols['MES ASSINADO']]).strip().upper() if row[cols.get('MES ASSINADO',11)] else '')
        mes_ass = MONTH_FIX.get(mes_ass, mes_ass)
        temp = (str(row[cols['TEMPERATURA']]).strip().upper() if row[cols.get('TEMPERATURA',12)] else 'INDEFINIDA')
        temp = {'NAN':'INDEFINIDA','NONE':'INDEFINIDA'}.get(temp, temp)
        uf = get_uf(regiao); cidade = clean_city(regiao)
        regiao_norm = f"{cidade} / {uf}" if uf!='N/D' else cidade
        records.append({
            'hunter': hunter, 'empresa': empresa, 'regiao': regiao_norm, 'uf': uf, 'cidade': cidade,
            'produto': str(produto).strip(), 'contrato': contrato, 'parceiro': parceiro,
            'honorarios': float(honor) if isinstance(honor,(int,float)) else 0,
            'faturamento': fat, 'credito': cred, 'status': status,
            'mes': mes, 'mes_assinado': mes_ass, 'temperatura': temp,
        })

    sh = wb['META']
    meta = []
    for row in sh.iter_rows(min_row=3, values_only=True):
        if row[1] is None: continue
        name = str(row[1]).strip().upper()
        name = HUNTER_FIX.get(name, name)
        name = 'NATHALIA' if name in ('NATALIA','NATHALIA') else name
        vals = {}
        for i,m in enumerate(MONTHS):
            v = row[2+i] if 2+i < len(row) else None
            vals[m] = (v*1000 if v else 0)
        meta.append({'hunter': name, 'meta': vals})

    sheets_info = {
        'Procurações Jan':   {'mes':'JAN', 'data_start':4, 'cols':{'empresa':2,'hunter':6,'proc':8}},
        'Procuraçoes Fev':   {'mes':'FEV', 'data_start':3, 'cols':{'empresa':2,'hunter':6,'proc':8}},
        'Procurações Mar':   {'mes':'MAR', 'data_start':3, 'cols':{'empresa':1,'hunter':5,'proc':7}},
        'Procurações Abril': {'mes':'ABR', 'data_start':2, 'cols':{'empresa':1,'hunter':5,'proc':6}},
        'procurações Maio':  {'mes':'MAI', 'data_start':2, 'cols':{'empresa':1,'hunter':5,'proc':6}},
    }
    procs = []
    for sn, info in sheets_info.items():
        if sn not in wb.sheetnames: continue
        sh = wb[sn]
        c = info['cols']
        for row in sh.iter_rows(min_row=info['data_start'], values_only=True):
            empresa = row[c['empresa']] if c['empresa']<len(row) else None
            if not empresa: continue
            hunter = row[c['hunter']] if c['hunter']<len(row) else None
            proc = row[c['proc']] if c['proc']<len(row) else None
            if not hunter and not proc: continue
            h = str(hunter).strip().upper() if hunter else 'N/D'
            h = HUNTER_FIX.get(h, h)
            procs.append({
                'mes': info['mes'],
                'empresa': str(empresa).strip(),
                'hunter': h,
                'tipo': str(proc).strip().upper() if proc else 'N/D',
            })

    tmpl_path = os.path.join(os.path.dirname(__file__), 'demo_template.html')
    with open(tmpl_path, 'r', encoding='utf-8') as f:
        html = f.read()
    html = html.replace('__DATE__', datetime.datetime.now().strftime('%d/%m/%Y %H:%M'))
    html = html.replace('__RECORDS__', json.dumps(records, ensure_ascii=False))
    html = html.replace('__META__', json.dumps(meta, ensure_ascii=False))
    html = html.replace('__PROCS__', json.dumps(procs, ensure_ascii=False))
    with open(OUT, 'w', encoding='utf-8') as f:
        f.write(html)
    size = os.path.getsize(OUT)
    print(f'DEMO gerada: {OUT} ({size:,} bytes)')
    print(f'  Forecast: {len(records)} | Meta: {len(meta)} | Procs: {len(procs)}')

if __name__ == '__main__':
    main()
