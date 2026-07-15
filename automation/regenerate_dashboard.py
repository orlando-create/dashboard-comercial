#!/usr/bin/env python3
import os, sys, json, re, datetime, subprocess, shutil, tempfile
import openpyxl

def _detect_sessions_root():
    base = '/sessions'
    if os.path.isdir(base):
        for d in os.listdir(base):
            mnt = os.path.join(base, d, 'mnt')
            if os.path.isdir(mnt) and os.path.isdir(os.path.join(mnt, 'Indicadores Time Comercial')):
                return os.path.join(base, d, 'mnt')
    return None

def _detect_windows_paths():
    candidates_indicadores = [
        r'C:\Users\orlan\OneDrive\Documentos\Claude\Projects\Indicadores Time Comercial',
    ]
    candidates_planejamento = [
        r'G:\.shortcut-targets-by-id\1Oe0H4_HJpjKQNma_pYUsvou4YE3tZckZ\Planejamento 2026',
        r'C:\Users\orlan\Meu Drive\COMERCIAL STRATEGICOS\Planejamento 2026',
        r'G:\Meu Drive\COMERCIAL STRATEGICOS\Planejamento 2026',
        r'G:\COMERCIAL STRATEGICOS\Planejamento 2026',
    ]
    ind_dir = next((p for p in candidates_indicadores if os.path.isdir(p)), None)
    pla_dir = next((p for p in candidates_planejamento if os.path.isdir(p)), None)
    return ind_dir, pla_dir

_MNT = _detect_sessions_root()
if _MNT is not None:
    _INDICADORES_DIR  = os.path.join(_MNT, 'Indicadores Time Comercial')
    _PLANEJAMENTO_DIR = os.path.join(_MNT, 'Planejamento 2026')
else:
    _ind, _pla = _detect_windows_paths()
    if _ind is None:
        raise RuntimeError('Pasta "Indicadores Time Comercial" nao encontrada. Verifique se o OneDrive esta sincronizado.')
    _INDICADORES_DIR  = _ind
    _PLANEJAMENTO_DIR = _pla or ''
    if not _PLANEJAMENTO_DIR:
        print('AVISO: pasta Planejamento 2026 nao encontrada no G: nem no Meu Drive.')

LAST_GOOD    = os.path.join(_INDICADORES_DIR, 'automation', 'forescast_last_good.xlsx')
SOURCE_XLSX  = LAST_GOOD
OUTPUT_HTML  = os.path.join(_INDICADORES_DIR, 'One_Page_Comercial_2026.html')
LOG_FILE     = os.path.join(_INDICADORES_DIR, 'automation', 'last_run.log')
TEMPLATE     = os.path.join(_INDICADORES_DIR, 'demo', 'Dashboard_Comercial_2026_v2.html')

UPS_ANUAL = {
  'RAFAEL': 2_000_000, 'HUDSON': 0, 'DIMAS': 923_000, 'HEITOR': 923_000,
  'CINTHIA': 0, 'BRUNA': 300_000, 'LEANDRO': 250_000,
  'ORLANDO': 923_000, 'ANDERSON': 0, 'RAUL': 0, 'CESAR ORGANICO': 2_000_000,
  'THIAGO': 0, 'KATARINY': 0, 'SILAS': 0, 'MARCELO': 0,
}
EXCLUIR_META = {'NATHALIA', 'NATALIA'}
MONTHS = ['JAN','FEV','MAR','ABR','MAI','JUN','JUL','AGO','SET','OUT','NOV','DEZ']

def parse_brl(v):
    if v is None: return 0.0
    if isinstance(v,(int,float)): return float(v)
    s = str(v).strip().replace('R$','').replace(' ','').replace('.','').replace(',','.')
    try: return float(s)
    except: return 0.0

def get_uf(r):
    if not r: return 'N/D'
    s = str(r).strip().upper()
    m = re.search(r'-\s*([A-Z]{2})\s*$', s)
    if m: return m.group(1)
    if s == 'SAO PAULO': return 'SP'
    return 'N/D'

def clean_city(r):
    if not r: return 'N/D'
    s = str(r).strip().upper()
    s = re.sub(r'\s*-\s*[A-Z]{2}\s*$','', s).strip()
    s = re.sub(r'\s+',' ', s)
    return s if s else 'N/D'

def log(msg):
    ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f'[{ts}] {msg}'
    print(line, flush=True)
    try:
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(line + '\n')
    except Exception:
        pass

def _rebuild_xlsx_from_raw(path):
    import struct, zlib as _zl, io as _io, zipfile as _zf, re as _re
    with open(path, 'rb') as f:
        data = f.read()
    locs = [m.start() for m in _re.finditer(b'PK\x03\x04', data)]
    dds  = [m.start() for m in _re.finditer(b'PK\x07\x08', data)]
    blocks = []
    for loc in locs:
        _, flag, comp, _, _, _, _, _, fl, el = struct.unpack_from('<HHHHHHIIHH', data, loc+4)
        ds = loc + 30 + fl + el
        ddp = next((d for d in dds if d >= ds), None)
        compressed = data[ds:ddp] if ddp else data[ds:]
        if ddp: dds = [d for d in dds if d != ddp]
        try:
            raw = _zl.decompress(compressed, -15) if comp == 8 else compressed
        except Exception:
            raw = b''
        blocks.append(raw)
    if len(blocks) < 10:
        raise RuntimeError(f'Reconstrucao falhou: {len(blocks)} blocos insuficientes')
    def bsearch(kw):
        kw_b = kw.encode() if isinstance(kw, str) else kw
        return next((i for i,b in enumerate(blocks) if b and kw_b in b), None)
    wb_idx = bsearch(b'<workbook')
    rl_idx = bsearch(b'worksheets/sheet1.xml')
    if wb_idx is None or rl_idx is None:
        raise RuntimeError(f'Reconstrucao falhou: workbook={wb_idx} rels={rl_idx}')
    sheet_blocks, rels_blocks = [], []
    i = 0
    while i < wb_idx:
        b = blocks[i]
        if b and b'<worksheet' in b[:300]:
            sheet_blocks.append(i)
            nxt = blocks[i+1] if i+1 < wb_idx else b''
            if nxt and b'Relationships' in nxt[:300]:
                rels_blocks.append(i+1); i += 2; continue
        i += 1
    drawing_blocks = [i for i,b in enumerate(blocks) if b and b'wsDr' in b[:200]]
    def fb(kw):
        kw_b = kw.encode() if isinstance(kw,str) else kw
        return next((i for i,b in enumerate(blocks) if b and kw_b in b[:600]), None)
    file_map = {}
    for j,di in enumerate(drawing_blocks):
        file_map[f'xl/drawings/drawing{j+1}.xml'] = blocks[di]
    for j,(si,ri) in enumerate(zip(sheet_blocks, rels_blocks)):
        n = j+1
        file_map[f'xl/worksheets/sheet{n}.xml']             = blocks[si]
        file_map[f'xl/worksheets/_rels/sheet{n}.xml.rels']  = blocks[ri]
    for key,kw in [('xl/theme/theme1.xml',b'<a:theme'),('xl/sharedStrings.xml',b'<sst '),
                    ('xl/styles.xml',b'<styleSheet'),('docProps/core.xml',b'cp:coreProperties'),
                    ('docProps/app.xml',b'<Properties ')]:
        idx = bsearch(kw)
        if idx is None and key == 'xl/sharedStrings.xml':
            idx = bsearch(b'<si>')
        if idx is not None: file_map[key] = blocks[idx]
    file_map['xl/workbook.xml']            = blocks[wb_idx]
    file_map['xl/_rels/workbook.xml.rels'] = blocks[rl_idx]
    ns = len(sheet_blocks)
    ct = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
          '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
          '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
          '<Default Extension="xml" ContentType="application/xml"/>',
          '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>']
    for n in range(1,ns+1):
        ct.append(f'<Override PartName="/xl/worksheets/sheet{n}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>')
    for n in range(1,len(drawing_blocks)+1):
        ct.append(f'<Override PartName="/xl/drawings/drawing{n}.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>')
    ct += ['<Override PartName="/xl/theme/theme1.xml" ContentType="application/vnd.openxmlformats-officedocument.theme+xml"/>',
           '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>',
           '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>',
           '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>',
           '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>',
           '</Types>']
    file_map['[Content_Types].xml'] = ''.join(ct).encode('utf-8')
    file_map['_rels/.rels'] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        '</Relationships>'
    ).encode('utf-8')
    buf = _io.BytesIO()
    with _zf.ZipFile(buf, 'w', _zf.ZIP_DEFLATED) as zf:
        for fname, content in file_map.items():
            if content: zf.writestr(fname, content)
    buf.seek(0)
    return buf


def _load_workbook_resilient(path):
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        log(f'AVISO: leitura direta falhou ({type(e).__name__}). Tentando zip -FF...')
    try:
        tmpdir  = tempfile.mkdtemp(prefix='xlsx_repair_')
        broken  = os.path.join(tmpdir, 'broken.xlsx')
        repaired = os.path.join(tmpdir, 'repaired.xlsx')
        shutil.copyfile(path, broken)
        subprocess.run(['zip', '-FF', broken, '--out', repaired],
                       input=b'y\n', capture_output=True, timeout=60)
        if os.path.exists(repaired):
            wb = openpyxl.load_workbook(repaired, data_only=True)
            log(f'Reparo zip -FF OK: {os.path.getsize(repaired):,} bytes')
            shutil.copyfile(repaired, LAST_GOOD)
            log(f'Copia last-good atualizada')
            return wb
    except Exception:
        pass
    log('zip -FF falhou. Tentando reconstrucao por conteudo...')
    try:
        buf = _rebuild_xlsx_from_raw(path)
        wb = openpyxl.load_workbook(buf, data_only=True)
        log(f'Reconstrucao OK: {len(wb.sheetnames)} abas')
        buf.seek(0)
        with open(LAST_GOOD, 'wb') as f:
            f.write(buf.read())
        log(f'Copia last-good atualizada')
        return wb
    except Exception as e:
        log(f'Reconstrucao falhou: {e}')
    if os.path.exists(LAST_GOOD):
        log(f'AVISO: usando ultima copia boa: {LAST_GOOD}')
        wb = openpyxl.load_workbook(LAST_GOOD, data_only=True)
        log(f'Last-good carregada OK: {len(wb.sheetnames)} abas')
        return wb
    raise RuntimeError('Todas as estrategias de leitura falharam e nao ha copia last-good')


def _pick_source():
    import glob as _glob, datetime
    candidates = []
    if os.path.isdir(_PLANEJAMENTO_DIR):
        for path in _glob.glob(os.path.join(_PLANEJAMENTO_DIR, '*.xlsx')):
            try:
                if not os.access(path, os.R_OK): continue
                mtime = os.path.getmtime(path)
                size  = os.path.getsize(path)
                if size < 10000: continue
                candidates.append((path, mtime))
            except Exception:
                pass
    candidates.sort(key=lambda x: -x[1])
    for path, mtime in candidates:
        ts = datetime.datetime.fromtimestamp(mtime).strftime('%d/%m/%Y %H:%M')
        log(f'Fonte encontrada: {os.path.basename(path)} | {ts}')
    if candidates:
        chosen = candidates[0][0]
        log(f'Usando (mais recente): {os.path.basename(chosen)}')
        return chosen
    if os.path.exists(LAST_GOOD):
        log('Nenhuma fonte em Planejamento 2026 acessivel. Usando last_good.')
        return LAST_GOOD
    log('ERRO: nenhum arquivo fonte encontrado.')
    return None


def main():
    log('=' * 60)
    log('Inicio da regeneracao')
    chosen = _pick_source()
    if chosen is None:
        log('ERRO: nenhum arquivo fonte disponivel')
        sys.exit(1)
    log(f'Lendo: {chosen}')
    wb = _load_workbook_resilient(chosen)

    sh = wb['FORECAST 2026']
    rows = list(sh.iter_rows(values_only=True))
    headers = list(rows[0])
    data_rows = rows[1:]
    cols = {h:i for i,h in enumerate(headers) if h}

    REGIAO_KEY = 'REGIÃO'  if 'REGIÃO'    in cols else 'REGIAO'
    HONOR_KEY  = 'HONORÁRIOS' if 'HONORÁRIOS' in cols else 'HONORARIOS'
    CRED_KEY   = 'CRÉDITO'  if 'CRÉDITO'   in cols else 'CREDITO'

    MES_NORM = {'ABRI':'ABR','MAIO':'MAI','JANEIRO':'JAN','FEVEREIRO':'FEV',
                'MARCO':'MAR','ABRIL':'ABR','JUNHO':'JUN',
                'JULHO':'JUL','AGOSTO':'AGO','SETEMBRO':'SET','OUTUBRO':'OUT',
                'NOVEMBRO':'NOV','DEZEMBRO':'DEZ','NAN':'','NONE':'',
                'JAN':'JAN','FEV':'FEV','MAR':'MAR','ABR':'ABR','MAI':'MAI',
                'JUN':'JUN','JUL':'JUL','AGO':'AGO','SET':'SET','OUT':'OUT',
                'NOV':'NOV','DEZ':'DEZ'}

    CI = {
        'HUNTER':   cols.get('HUNTER', cols.get('HUBTER', 0)),
        'EMPRESA':  cols.get('EMPRESA',  1),
        'REGIAO':   cols.get(REGIAO_KEY, 2),
        'PRODUTO':  cols.get('PRODUTO',  3),
        'CONTRATO': cols.get('CONTRATO', 4),
        'PARCEIRO': cols.get('PARCEIRO', 5),
        'CREDITO':  cols.get(CRED_KEY,   6),
        'HONOR':    cols.get(HONOR_KEY,  7),
        'FAT':      cols.get('FATURAMENTO', 8),
        'STATUS':   cols.get('STATUS',   9),
        'MES':      cols.get('MES',      10),
        'MES_ASS':  cols.get('MES ASSINADO', 11),
        'ORIGEM':   cols.get('Origem',   12),
        'PROC':     cols.get('PROCURACAO CADASTRADA ?', cols.get('PROCURAÇÃO CADASTRADA ?', cols.get('PROCURACAO CADASTRADA?', 14))),
        'TEMP':     cols.get('TEMPERATURA', 14),
    }

    HUNTER_OVERRIDE = {'SUPER DABARRA LTDA': 'ANDERSON'}

    records = []
    for row in data_rows:
        if not row or len(row) <= CI['HUNTER'] or row[CI['HUNTER']] is None: continue
        hunter = str(row[CI['HUNTER']]).strip().upper()
        hunter = {'NATALIA':'NATHALIA','THIAGO ':'THIAGO'}.get(hunter, hunter)
        empresa = str(row[CI['EMPRESA']]).strip() if CI['EMPRESA'] < len(row) and row[CI['EMPRESA']] else ''
        if empresa in HUNTER_OVERRIDE: hunter = HUNTER_OVERRIDE[empresa]
        if not empresa: continue
        regiao   = row[CI['REGIAO']]   if CI['REGIAO']   < len(row) else None
        produto  = row[CI['PRODUTO']]  if CI['PRODUTO']  < len(row) else None
        produto  = produto or 'OUTROS'
        contrato = str(row[CI['CONTRATO']]).strip().upper() if CI['CONTRATO'] < len(row) and row[CI['CONTRATO']] else 'NOVO'
        contrato = 'NOVO' if contrato in ('','NAN','NONE') else contrato
        parceiro = str(row[CI['PARCEIRO']]).strip().upper() if CI['PARCEIRO'] < len(row) and row[CI['PARCEIRO']] else 'NAO'
        honor    = row[CI['HONOR']] if CI['HONOR'] < len(row) else 0
        honor    = honor or 0
        fat      = parse_brl(row[CI['FAT']]    if CI['FAT']    < len(row) else None)
        cred     = parse_brl(row[CI['CREDITO']] if CI['CREDITO'] < len(row) else None)
        status   = str(row[CI['STATUS']]).strip().upper()  if CI['STATUS']  < len(row) and row[CI['STATUS']]  else ''
        mes      = MES_NORM.get(str(row[CI['MES']]).strip().upper()     if CI['MES']     < len(row) and row[CI['MES']]     else '', '')
        mes_ass  = MES_NORM.get(str(row[CI['MES_ASS']]).strip().upper() if CI['MES_ASS'] < len(row) and row[CI['MES_ASS']] else '', '')
        temp     = str(row[CI['TEMP']]).strip().upper() if CI['TEMP'] < len(row) and row[CI['TEMP']] else 'INDEFINIDA'
        temp     = {'NAN':'INDEFINIDA','NONE':'INDEFINIDA'}.get(temp, temp)
        uf = get_uf(regiao); cidade = clean_city(regiao)
        origem_raw = row[CI['ORIGEM']] if CI['ORIGEM'] < len(row) else None
        proc_raw   = row[CI['PROC']]   if CI['PROC']   < len(row) else None
        records.append({
            'hunter': hunter, 'empresa': empresa,
            'regiao': f"{cidade} / {uf}" if uf!='N/D' else cidade,
            'uf': uf, 'cidade': cidade,
            'produto': str(produto).strip(), 'contrato': contrato, 'parceiro': parceiro,
            'honorarios': float(honor) if isinstance(honor,(int,float)) else 0,
            'faturamento': fat, 'credito': cred, 'status': status,
            'mes': mes, 'mes_assinado': mes_ass, 'temperatura': temp,
            'origem': str(origem_raw).strip() if origem_raw else '',
            'proc_cadastrada': str(proc_raw).strip().upper() if proc_raw else '',
        })

    log(f'Forecast: {len(records)} registros lidos')

    sh = wb['META']
    meta, excluidos = [], []
    for row in sh.iter_rows(min_row=3, values_only=True):
        if row[1] is None: continue
        name = str(row[1]).strip().upper()
        if name in ('TOTAL', 'TOTAL ANO', 'TOTAIS'): continue  # linha de resumo, nao e um hunter
        name = 'NATHALIA' if name in ('NATALIA','NATHALIA') else name
        if name in EXCLUIR_META: excluidos.append(name); continue
        vals = {}
        for i,m in enumerate(MONTHS):
            v = row[2+i] if 2+i < len(row) else None
            if v is None or v == '':
                vals[m] = 0
            else:
                v = float(v)
                vals[m] = v*1000 if v < 5000 else v
        meta.append({'hunter': name, 'meta': vals})

    existing = set(m['hunter'] for m in meta)
    for h in UPS_ANUAL:
        if h in EXCLUIR_META: continue
        if h not in existing:
            meta.append({'hunter': h, 'meta': {m: 0 for m in MONTHS}})
    for m in meta:
        anual = UPS_ANUAL.get(m['hunter'], 0)
        m['meta_upsell'] = {mo: round(anual/12, 2) for mo in MONTHS}

    total_novo = sum(sum(m['meta'].values()) for m in meta)
    total_ups  = sum(sum(m['meta_upsell'].values()) for m in meta)
    log(f'Meta: {len(meta)} hunters | NOVO R$ {total_novo:,.0f} | UPSELL R$ {total_ups:,.0f} | TOTAL R$ {(total_novo+total_ups):,.0f}')
    if excluidos: log(f'Excluidos da meta: {", ".join(sorted(set(excluidos)))}')

    sheets_info = {
        'Procuracoes Jan':   {'mes':'JAN','data_start':4,'cols':{'empresa':2,'hunter':6,'proc':8}},
        'Procuracoes Fev':   {'mes':'FEV','data_start':3,'cols':{'empresa':2,'hunter':6,'proc':8}},
        'Procuracoes Mar':   {'mes':'MAR','data_start':3,'cols':{'empresa':1,'hunter':5,'proc':7}},
        'Procuracoes Abril': {'mes':'ABR','data_start':2,'cols':{'empresa':1,'hunter':5,'proc':6}},
        'procuracoes Maio':  {'mes':'MAI','data_start':2,'cols':{'empresa':1,'hunter':5,'proc':6}},
        'Procuracoes de Junho': {'mes':'JUN','data_start':2,'cols':{'empresa':1,'hunter':5,'proc':6}},
        'Procuracoes de Julho': {'mes':'JUL','data_start':2,'cols':{'empresa':1,'hunter':5,'proc':6}},
    }
    sheets_info_accented = {
        'Procuracoes Jan':   'Procurações Jan',
        'Procuracoes Fev':   'Procuraçoes Fev',
        'Procuracoes Mar':   'Procurações Mar',
        'Procuracoes Abril': 'Procurações Abril',
        'procuracoes Maio':  'procurações Maio',
        'Procuracoes de Junho': 'Procurações de Junho',
        'Procuracoes de Julho': 'Procurações de Julho',
    }
    procs = []
    for sn, info in sheets_info.items():
        actual_sn = sn
        if sn not in wb.sheetnames and sheets_info_accented.get(sn) in wb.sheetnames:
            actual_sn = sheets_info_accented[sn]
        if actual_sn not in wb.sheetnames:
            log(f'AVISO: aba {sn} nao encontrada')
            continue
        sh = wb[actual_sn]; c = info['cols']
        for row in sh.iter_rows(min_row=info['data_start'], values_only=True):
            empresa = row[c['empresa']] if c['empresa']<len(row) else None
            if not empresa: continue
            hunter = row[c['hunter']] if c['hunter']<len(row) else None
            proc   = row[c['proc']]   if c['proc']<len(row)   else None
            if not hunter and not proc: continue
            h = str(hunter).strip().upper() if hunter else 'N/D'
            h = {'NATALIA':'NATHALIA','THIAGO ':'THIAGO'}.get(h, h)
            procs.append({'mes': info['mes'], 'empresa': str(empresa).strip(),
                          'hunter': h, 'tipo': str(proc).strip().upper() if proc else 'N/D'})
    log(f'Procuracoes: {len(procs)} registros')

    sh_fc    = wb['FORECAST 2026']
    rows_fc  = list(sh_fc.iter_rows(values_only=True))
    cols_fc  = {h:i for i,h in enumerate(rows_fc[0]) if h}
    origem_col = CI['ORIGEM']
    proc_col   = CI['PROC']

    by_hunter_apas  = {}
    empresas_apas   = set()
    emp_apas_proc   = {}
    total_fat_apas  = 0.0
    total_cred_apas = 0.0
    pendentes_apas  = 0
    linhas_apas     = 0

    for row in rows_fc[1:]:
        if origem_col is None: break
        ov = row[origem_col] if origem_col < len(row) else None
        if not ov or 'APAS' not in str(ov).upper(): continue
        linhas_apas += 1
        hunter  = str(row[CI['HUNTER']]).strip().upper() if CI['HUNTER'] < len(row) and row[CI['HUNTER']] else 'N/D'
        hunter  = {'NATALIA':'NATHALIA','THIAGO ':'THIAGO'}.get(hunter, hunter)
        empresa = str(row[CI['EMPRESA']]).strip() if CI['EMPRESA'] < len(row) and row[CI['EMPRESA']] else ''
        if empresa in HUNTER_OVERRIDE: hunter = HUNTER_OVERRIDE[empresa]
        fat     = parse_brl(row[CI['FAT']]     if CI['FAT']     < len(row) else None)
        cred    = parse_brl(row[CI['CREDITO']] if CI['CREDITO'] < len(row) else None)
        proc_v  = row[proc_col] if proc_col is not None and proc_col < len(row) else None
        tem_proc = str(proc_v).strip().upper() == 'SIM' if proc_v else False
        total_fat_apas  += fat
        total_cred_apas += cred
        if fat == 0: pendentes_apas += 1
        if empresa:
            empresas_apas.add(empresa)
            eu = empresa.strip().upper()
            emp_apas_proc[eu] = emp_apas_proc.get(eu, False) or tem_proc
        if hunter not in by_hunter_apas:
            by_hunter_apas[hunter] = {'fat':0.0,'cred':0.0,'empresas':set(),'empresas_proc':set(),'pendentes':0}
        by_hunter_apas[hunter]['fat']  += fat
        by_hunter_apas[hunter]['cred'] += cred
        if empresa:
            by_hunter_apas[hunter]['empresas'].add(empresa)
            if tem_proc: by_hunter_apas[hunter]['empresas_proc'].add(empresa)
        if fat == 0: by_hunter_apas[hunter]['pendentes'] += 1

    empresas_proc_mensal = set()
    for sn, info in sheets_info.items():
        actual_sn = sn
        if sn not in wb.sheetnames and sheets_info_accented.get(sn) in wb.sheetnames:
            actual_sn = sheets_info_accented[sn]
        if actual_sn not in wb.sheetnames:
            continue
        sh_p = wb[actual_sn]; c = info['cols']
        for row in sh_p.iter_rows(min_row=info['data_start'], values_only=True):
            emp = row[c['empresa']] if c['empresa'] < len(row) else None
            if emp:
                empresas_proc_mensal.add(str(emp).strip().upper())
    for eu in list(emp_apas_proc.keys()):
        emp_apas_proc[eu] = emp_apas_proc[eu] or (eu in empresas_proc_mensal)
    for eu in set(e.strip().upper() for e in empresas_apas):
        if eu not in emp_apas_proc:
            emp_apas_proc[eu] = eu in empresas_proc_mensal
    apas_proc_sim = sum(1 for v in emp_apas_proc.values() if v)
    apas_proc_nao = sum(1 for v in emp_apas_proc.values() if not v)
    log(f'APAS x Procuracoes: {apas_proc_sim} com cadastro | {apas_proc_nao} sem cadastro (de {len(emp_apas_proc)} empresas unicas)')

    apas_by_hunter_list = sorted(
        [{'hunter': h, 'fat': v['fat'], 'cred': v['cred'],
          'empresas': len(v['empresas']), 'pendentes': v['pendentes'],
          'proc_sim': len(v['empresas_proc']),
          'proc_nao': len(v['empresas']) - len(v['empresas_proc']),
          } for h, v in by_hunter_apas.items()],
        key=lambda x: -x['fat']
    )

    apas_data = {
        'total_fat':  round(total_fat_apas, 2),
        'total_cred': round(total_cred_apas, 2),
        'empresas':   len(empresas_apas),
        'linhas':     linhas_apas,
        'pendentes':  pendentes_apas,
        'proc_sim':   apas_proc_sim,
        'proc_nao':   apas_proc_nao,
        'by_hunter':  apas_by_hunter_list,
    }
    log(f'APAS Show: {linhas_apas} linhas | {len(empresas_apas)} empresas | Fat R$ {total_fat_apas:,.0f} | Cred R$ {total_cred_apas:,.0f} | {pendentes_apas} pendentes')

    records_d = []
    for r in records:
        records_d.append({
            'hunter':    r['hunter'],
            'empresa':   r['empresa'],
            'uf':        r['uf'],
            'cidade':    r.get('cidade', ''),
            'produto':   r['produto'],
            'contrato':  r['contrato'],
            'parceiro':  r.get('parceiro', ''),
            'credito':   r.get('credito', 0),
            'honorarios': r.get('honorarios', 0),
            'fat':       r['faturamento'],
            'status':    r['status'],
            'mes':       r['mes'],
            'mes_ass':   r.get('mes_assinado', ''),
            'temp':      r.get('temperatura', ''),
            'origem':    r.get('origem', ''),
            'is_apas':   any(a in str(r.get('origem', '')).strip().upper() for a in ('APAS', 'APAS SHOW')),
            'proc_ok':   bool(str(r.get('proc_cadastrada', '')).strip()),
        })
    meta_monthly = [sum(h['meta'].get(m, 0) for h in meta) for m in MONTHS]
    meta_hunters  = {h['hunter']: [h['meta'].get(m, 0) for m in MONTHS] for h in meta}
    ups_monthly   = [sum(h.get('meta_upsell', {}).get(m, 0) for h in meta) for m in MONTHS]
    ups_hunters   = {h['hunter']: [h.get('meta_upsell', {}).get(m, 0) for m in MONTHS] for h in meta}
    d_data = {
        'records':      records_d,
        'meta_monthly': meta_monthly,
        'meta_hunters': meta_hunters,
        'ups_monthly':  ups_monthly,
        'ups_hunters':  ups_hunters,
        'proc_records': procs,
    }

    with open(TEMPLATE, 'r', encoding='utf-8') as f:
        html = f.read()

    # Guarda de integridade: o OneDrive/sandbox às vezes serve uma cópia local
    # parcialmente sincronizada (cache antigo, arquivo cortado no meio de uma
    # linha). Isso já aconteceu e gerou um One_Page_Comercial_2026.html quebrado
    # (JS com erro de sintaxe -> dashboard inteiro em branco). Antes de gerar
    # qualquer coisa, valida que o template lido está de fato completo.
    _tail_ok = html.rstrip().endswith('</html>')
    _n_open  = len(re.findall(r'<script[^>]*>', html))
    _n_close = len(re.findall(r'</script>', html))
    if not _tail_ok or _n_open != _n_close or len(html) < 80000:
        raise RuntimeError(
            f'Template "{TEMPLATE}" parece truncado/incompleto '
            f'(tamanho={len(html)} bytes, termina em </html>={_tail_ok}, '
            f'scripts abertos={_n_open} fechados={_n_close}). '
            'Abortando para não publicar um dashboard quebrado. '
            'Isso costuma ser um cache local do OneDrive desatualizado — '
            'tente novamente ou force uma releitura completa do arquivo antes '
            'de rodar a automação de novo.'
        )

    html = html.replace('REPLACE_DATE',    datetime.datetime.now().strftime('%d/%m/%Y %H:%M'))
    html = html.replace('REPLACE_D_JSON',  json.dumps(d_data, ensure_ascii=False))
    html = html.replace('REPLACE_DATA_JSON',  json.dumps(records,   ensure_ascii=False))
    html = html.replace('REPLACE_META_JSON',  json.dumps(meta,      ensure_ascii=False))
    html = html.replace('REPLACE_PROCS_JSON', json.dumps(procs,     ensure_ascii=False))
    html = html.replace('REPLACE_APAS_JSON',  json.dumps(apas_data, ensure_ascii=False))

    html = html.replace(
        "if(F.mes&&(r.mes_ass||r.mes)!==F.mes)return false;",
        "if(F.mes&&r.mes_ass!==F.mes)return false;"
    )
    html = html.replace(
        "ass.forEach(r=>{const m=r.mes_ass||r.mes;if(bym[m])bym[m].novo+=r.fat;});",
        "ass.forEach(r=>{const m=r.mes_ass;if(m&&bym[m])bym[m].novo+=r.fat;});"
    )
    html = html.replace(
        "base.forEach(r=>{const m=r.mes_ass||r.mes;if(bym[m])bym[m].base+=r.fat;});",
        "base.forEach(r=>{const m=r.mes_ass;if(m&&bym[m])bym[m].base+=r.fat;});"
    )

    import re as _re2
    html = _re2.sub(
        r'(sa\.map\(\(\[e,v\],i\)=>`[^`]+`);\}\)\.join',
        lambda m: m.group(1) + ').join',
        html
    )

    import re as _re
    has_complete_boot = bool(_re.search(r'function boot\s*\(\s*\)\s*\{', html))

    if has_complete_boot:
        log('Template com boot() completo — sem injecao adicional')
    else:
        log('Template sem boot() — injetando boot de seguranca')
        BOOT_SUFFIX = (
            "\n<script>\n"
            "/* boot */\n"
            "(function(){\n"
            "  function _boot(){\n"
            "    try{\n"
            "      if(typeof buildFilters==='function') buildFilters();\n"
            "      if(typeof renderAll==='function') renderAll();\n"
            "    }catch(e){\n"
            "      var el=document.getElementById('rec-count');\n"
            "      if(el) el.textContent='Erro: '+e.message;\n"
            "      console.error('[boot]',e);\n"
            "    }\n"
            "  }\n"
            "  if(document.readyState==='loading'){\n"
            "    document.addEventListener('DOMContentLoaded',_boot);\n"
            "  } else { _boot(); }\n"
            "})();\n"
            "</script>\n"
            "</body>\n</html>"
        )
        for closing in ['</body>\n</html>', '</body></html>', '</html>']:
            if html.rstrip().endswith(closing.strip()):
                html = html.rstrip()[:-len(closing.strip())]
                break
        boot_match = _re.search(r'\nfunction boot\(\)\{[^<]*$', html, _re.DOTALL)
        if boot_match:
            log('AVISO: function boot() incompleta removida (truncamento OneDrive)')
            html = html[:boot_match.start()]
        n_open  = len(_re.findall(r'<script[^>]*>', html))
        n_close = len(_re.findall(r'</script>', html))
        if n_open > n_close:
            log(f'AVISO: {n_open - n_close} script(s) sem fechar — fechando antes do boot')
            html = html.rstrip() + '\n</script>\n'
        html = html + BOOT_SUFFIX

    tmp_html = os.path.join(tempfile.gettempdir(), 'One_Page_tmp_{}.html'.format(os.getpid()))
    with open(tmp_html, 'w', encoding='utf-8') as f:
        f.write(html)
    shutil.copy2(tmp_html, OUTPUT_HTML)
    os.remove(tmp_html)

    log(f'HTML gerado: {OUTPUT_HTML} ({os.path.getsize(OUTPUT_HTML):,} bytes)')
    print(f'__STATS__: forecast={len(records)} meta={len(meta)} procs={len(procs)} html_size={os.path.getsize(OUTPUT_HTML)}')

if __name__ == '__main__':
    main()
