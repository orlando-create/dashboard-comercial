"""
EXTRAIR DADOS PARA ARTIFACT — Indicadores Time Comercial
=========================================================
Lê a planilha Excel e imprime o JSON de dados para o artifact do dashboard.
Usado pela tarefa agendada para atualizar o dashboard online automaticamente.

Uso: python extrair_dados_artifact.py
"""

import openpyxl, json, os, sys
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, 'Dashboard_Comercial_2026.xlsx')

def parse_regiao(praca):
    if not praca: return 'OUTROS', 'OUTROS', 'OUTROS'
    p = str(praca).strip()
    mapping = {
        'São Paulo e Santos':   ('São Paulo e Santos',       'SP', 'SAO PAULO'),
        'Campinas':             ('Campinas',                 'SP', 'CAMPINAS'),
        'Fortaleza':            ('Fortaleza',                'CE', 'FORTALEZA'),
        'Curitiba':             ('Curitiba',                 'PR', 'CURITIBA'),
        'Goiânia':              ('Goiânia',                  'GO', 'GOIANIA'),
        'Goiania':              ('Goiânia',                  'GO', 'GOIANIA'),
        'Minas Gerais':         ('Minas Gerais',             'MG', 'BELO HORIZONTE'),
        'Rio de Janeiro':       ('Rio de Janeiro',           'RJ', 'RIO DE JANEIRO'),
        'Ribeirão Preto':       ('Ribeirão Preto',           'SP', 'RIBEIRAO PRETO'),
        'São José do Rio Preto':('São José do Rio Preto',    'SP', 'SJR PRETO'),
        'Marília e região':     ('Marília e região',         'SP', 'MARILIA'),
        'São Paulo - Matriz':   ('São Paulo - Matriz',       'SP', 'SAO PAULO'),
        'Teresina e MS':        ('Teresina e MS',            'PI/MS', 'TERESINA'),
        'Recife':               ('Recife',                   'PE', 'RECIFE'),
        'Matriz':               ('Matriz',                   'SP', 'SAO PAULO'),
        'SJC e ABC':            ('SJC e ABC',                'SP', 'SJC'),
    }
    for k, v in mapping.items():
        if k.lower() in p.lower():
            return v
    return (p, 'OUTROS', p.upper())

def safe_float(v, default=0.0):
    try:
        return round(float(v), 4) if v is not None else default
    except (ValueError, TypeError):
        return default

def safe_str(v, default=''):
    return str(v).strip() if v is not None else default

# ── Leitura do Excel ─────────────────────────────────────────────────────────
if not os.path.exists(EXCEL_PATH):
    print(json.dumps({"error": f"Arquivo não encontrado: {EXCEL_PATH}"}))
    sys.exit(1)

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# Detecta aba
sheet_name = None
for candidate in ['Base de Dados', 'BASE DE DADOS', 'Dados', 'Sheet1', 'Plan1']:
    if candidate in wb.sheetnames:
        sheet_name = candidate
        break
if not sheet_name:
    sheet_name = wb.sheetnames[0]

ws = wb[sheet_name]

# Mapear cabeçalhos
headers = {}
for c in range(1, ws.max_column + 1):
    val = ws.cell(1, c).value
    if val:
        headers[str(val).strip()] = c - 1

def col(row, name, default=None):
    idx = headers.get(name)
    if idx is None:
        return default
    v = row[idx]
    return v if v is not None else default

# ── Processar linhas ──────────────────────────────────────────────────────────
rows_out = []
hunters_seen = set()
praca_map = {}
uf_map = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    hunter = col(row, 'HUNTER')
    if not hunter:
        continue
    hunter = safe_str(hunter)

    praca_raw = col(row, 'PRAÇA') or col(row, 'PRACA') or ''
    praca_label, uf, _ = parse_regiao(praca_raw)

    faturamento = safe_float(col(row, 'FATURAMENTO'))
    status      = safe_str(col(row, 'STATUS'))
    mes         = safe_str(col(row, 'MÊS') or col(row, 'MES'))
    contrato    = safe_str(col(row, 'CONTRATO'), 'NOVO')
    categoria   = safe_str(col(row, 'PRODUTO') or col(row, 'CATEGORIA') or col(row, 'PRODUCT'))

    rows_out.append({
        "h":  hunter,
        "pr": praca_label,
        "uf": uf,
        "c":  categoria,
        "f":  faturamento,
        "s":  status,
        "m":  mes,
        "t":  contrato,
    })

    hunters_seen.add(hunter)
    if hunter not in praca_map and praca_label:
        praca_map[hunter] = praca_label
    if hunter not in uf_map and uf:
        uf_map[hunter] = uf

# ── Metas ─────────────────────────────────────────────────────────────────────
metas = {}
meta_candidates = ['Metas', 'METAS', 'Meta', 'META']
for candidate in meta_candidates:
    if candidate in wb.sheetnames:
        ws_meta = wb[candidate]
        meta_headers = {}
        for c in range(1, ws_meta.max_column + 1):
            v = ws_meta.cell(1, c).value
            if v:
                meta_headers[str(v).strip().upper()] = c - 1
        for row in ws_meta.iter_rows(min_row=2, values_only=True):
            h = row[0]
            if not h:
                continue
            h = str(h).strip()
            vals = [safe_float(row[i]) if i < len(row) else 0 for i in range(1, 13)]
            metas[h] = vals
        break

# Metas padrão se aba não encontrada (mantém os valores anteriores)
DEFAULT_METAS = {
    "HUDSON":    [168000,252000,336000,336000,252000,336000,504000,378000,504000,504000,378000,252000],
    "DIMAS":     [128000,192000,256000,256000,192000,256000,384000,288000,384000,384000,288000,192000],
    "HEITOR":    [128000,192000,256000,256000,192000,256000,384000,288000,384000,384000,288000,192000],
    "CINTHIA":   [128000,192000,256000,256000,192000,256000,384000,288000,384000,384000,288000,192000],
    "BRUNA":     [128000,192000,256000,256000,192000,256000,384000,288000,384000,384000,288000,192000],
    "LEANDRO":   [128000,192000,256000,256000,192000,256000,384000,288000,384000,384000,288000,192000],
    "NATHALIA":  [128000,192000,256000,256000,192000,256000,384000,288000,384000,384000,288000,192000],
    "ORLANDO":   [168000,252000,336000,336000,252000,336000,504000,378000,504000,504000,378000,252000],
    "RAUL":      [0,0,264000,198000,264000,264000,198000,132000,264000,198000,264000,128000],
    "ANDERSON":  [0,0,0,100000,100000,100000,300000,225000,300000,300000,225000,150000],
    "FORTALEZA 2":[0,0,0,0,0,0,300000,225000,300000,300000,225000,150000],
    "KATARINI":  [0,0,0,100000,100000,100000,264000,198000,264000,264000,198000,132000],
    "THIAGO":    [0,0,0,100000,100000,100000,264000,198000,264000,264000,198000,132000],
    "PIAUI 2":   [0,0,0,0,0,0,264000,198000,264000,264000,198000,132000],
    "SILAS":     [0,0,0,100000,100000,100000,264000,198000,264000,264000,198000,132000],
}
if not metas:
    metas = DEFAULT_METAS

# Metas padrão adicionadas para manter compatibilidade
DEFAULT_PRACA = {
    "BRUNA":"São Paulo e Santos","CINTHIA":"Fortaleza","ANDERSON":"Fortaleza",
    "DIMAS":"Teresina e MS","THIAGO":"Recife","KATARINI":"Recife",
    "HEITOR":"Campinas","LEANDRO":"Marília e região","RAUL":"SJC e ABC",
    "NATHALIA":"São José do Rio Preto","HUDSON":"São Paulo - Matriz",
    "ORLANDO":"São Paulo - Matriz","ARTHUR":"Matriz","SILAS":"Fortaleza",
    "RAFAEL":"São Paulo - Matriz","FORTALEZA 2":"Fortaleza","PIAUI 2":"Teresina"
}
DEFAULT_UF = {
    "BRUNA":"SP","CINTHIA":"CE","ANDERSON":"CE","DIMAS":"PI/MS","THIAGO":"PE",
    "KATARINI":"PE","HEITOR":"SP","LEANDRO":"SP","RAUL":"SP","NATHALIA":"SP",
    "HUDSON":"SP","ORLANDO":"SP","ARTHUR":"SP","SILAS":"CE","RAFAEL":"SP",
    "FORTALEZA 2":"CE","PIAUI 2":"PI"
}

# Complementar com defaults para hunters sem praça/uf mapeados
for h, v in DEFAULT_PRACA.items():
    if h not in praca_map:
        praca_map[h] = v
for h, v in DEFAULT_UF.items():
    if h not in uf_map:
        uf_map[h] = v

result = {
    "rows":      rows_out,
    "metas":     metas,
    "praca_map": praca_map,
    "uf_map":    uf_map,
    "updated_at": datetime.now().strftime('%d/%m/%Y %H:%M'),
    "total_rows": len(rows_out),
}

print(json.dumps(result, ensure_ascii=False, separators=(',', ':')))
